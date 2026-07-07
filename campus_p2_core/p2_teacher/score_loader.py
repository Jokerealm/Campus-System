from __future__ import annotations

from io import StringIO
from pathlib import Path
import csv

from openpyxl import load_workbook

from campus_p2_core.contracts.p2 import ScoreRecord


HEADER_ALIASES = {
    "question_no": ["题号", "小题号", "试题号", "question_no", "question", "no"],
    "full_score": ["满分", "分值", "题目满分", "full_score", "score", "points"],
    "avg_score": ["平均分", "班级平均分", "均分", "avg_score", "average", "mean"],
    "score_rate": ["得分率", "平均得分率", "正确率", "score_rate", "rate"],
    "sample_count": ["人数", "样本量", "sample_count", "count"],
}


def load_score_records(path: str | Path) -> list[ScoreRecord]:
    score_path = Path(path)
    suffix = score_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(score_path, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        return _parse_rows(rows)
    if suffix in {".csv", ".txt"}:
        text = score_path.read_text(encoding="utf-8-sig")
        rows = [row for row in csv.reader(StringIO(text))]
        return _parse_rows(rows)
    raise ValueError(f"Unsupported score file type: {score_path.suffix}")


def _parse_rows(rows: list[list[object]]) -> list[ScoreRecord]:
    if not rows:
        return []
    header_index, columns = _find_header(rows)
    missing = [field for field in ["question_no", "full_score"] if field not in columns]
    if "avg_score" not in columns and "score_rate" not in columns:
        missing.append("avg_score")
    if missing:
        raise ValueError(f"Missing score columns: {', '.join(missing)}")

    records: list[ScoreRecord] = []
    for row in rows[header_index + 1 :]:
        if not any(cell not in (None, "") for cell in row):
            continue
        question_no = str(_cell(row, columns["question_no"]) or "").strip()
        full_score = _to_float(_cell(row, columns["full_score"]))
        avg_score = _to_float(_cell(row, columns["avg_score"])) if "avg_score" in columns else None
        if avg_score is None and "score_rate" in columns and full_score is not None:
            rate = _score_rate_to_fraction(_cell(row, columns["score_rate"]))
            avg_score = full_score * rate if rate is not None else None
        sample_count = None
        if "sample_count" in columns:
            sample = _to_float(_cell(row, columns["sample_count"]))
            sample_count = int(sample) if sample is not None else None
        warnings = []
        if avg_score is not None and full_score is not None and avg_score > full_score:
            warnings.append("平均分超过满分")
        if question_no and full_score is not None and avg_score is not None:
            records.append(
                ScoreRecord(
                    question_no=question_no,
                    full_score=full_score,
                    avg_score=avg_score,
                    sample_count=sample_count,
                    warnings=warnings,
                )
            )
    return records


def _find_header(rows: list[list[object]]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows[:8]):
        columns = _detect_columns(row)
        if {"question_no", "full_score"}.issubset(columns) and ("avg_score" in columns or "score_rate" in columns):
            return index, columns
    return 0, _detect_columns(rows[0])


def _detect_columns(headers: list[object]) -> dict[str, int]:
    normalized = [_normalize_header(header) for header in headers]
    detected: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        alias_set = {_normalize_header(alias) for alias in aliases}
        for index, header in enumerate(normalized):
            if header in alias_set:
                detected[field] = index
                break
    return detected


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "")


def _cell(row: list[object], index: int) -> object | None:
    return row[index] if index < len(row) else None


def _to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        cleaned = value.strip().replace("%", "")
        if not cleaned:
            return None
        number = float(cleaned)
        return number / 100 if "%" in value else number
    return float(value)


def _score_rate_to_fraction(value: object) -> float | None:
    rate = _to_float(value)
    if rate is None:
        return None
    if rate > 1:
        rate = rate / 100
    return max(0.0, min(rate, 1.0))
