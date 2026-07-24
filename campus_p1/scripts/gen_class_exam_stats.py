"""生成 3 份试卷的班级阅卷统计 Excel（每题一行，聚合视图）。

设计原则：
- 50 人班级作答数据来自 (mock) 抽样分布，再聚合出每题统计
- 错题率随题号递增（与题目难度梯度一致）
- 选择题错题率最低，填空中等，解答题最高
- 典型错误从题目的 qwen_analysis.knowledge_points 反推一类高频错误描述
"""
from __future__ import annotations

import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PAPERS_DIR = Path("/home/yjk/Guan/word_cutter_system/web_app/backend/papers")

# 50 人班级、种子固定，便于复现
random.seed(42)
N_STUDENTS = 50

HEADER = [
    "题号", "题型", "分值",
    "班级均分", "最高分", "最低分", "标准差",
    "错题率", "零分率", "满分率",
    "典型错误", "难度等级",
]


def _classify_difficulty(q_no: int, q_type: str, total: int) -> str:
    """根据题号位置粗略判难度。后置题难。"""
    ratio = q_no / total
    if ratio <= 0.4:
        base = "易"
    elif ratio <= 0.75:
        base = "中"
    else:
        base = "难"
    if q_type == "single_choice":
        return {"易": "易", "中": "中", "难": "中"}[base]
    if q_type == "blank":
        return {"易": "易", "中": "中", "难": "难"}[base]
    return base  # solution


def _target_error_rate(q_type: str, difficulty: str) -> float:
    table = {
        "single_choice": {"易": 0.10, "中": 0.30, "难": 0.55},
        "blank":         {"易": 0.15, "中": 0.40, "难": 0.65},
        "solution":      {"易": 0.12, "中": 0.35, "难": 0.60},
    }
    return table[q_type][difficulty]


def _gen_student_scores(n: int, full_score: int, err_rate: float, difficulty: str) -> list[int]:
    """根据目标错题率生成每位同学此题得分（0~full_score）。"""
    scores = []
    zero_rate = max(0.02, err_rate * 0.4)         # 错题人数里约 40% 拿零分
    full_rate = max(0.0, 0.6 - err_rate)          # 简单题满分率高
    near_full = max(0.0, (1 - err_rate) - full_rate)
    partial = max(0.0, 1 - zero_rate - full_rate - near_full)

    for _ in range(n):
        r = random.random()
        if r < zero_rate:
            scores.append(0)
        elif r < zero_rate + full_rate:
            scores.append(full_score)
        elif r < zero_rate + full_rate + near_full:
            scores.append(full_score - 1 if full_score > 1 else full_score)
        else:
            # 部分得分：35%~85% of full_score
            frac = random.uniform(0.35, 0.85)
            scores.append(int(round(full_score * frac)))
    return scores


def _typical_error(q: dict) -> str:
    """从 qwen_analysis.knowledge_points + question_type 反推一句中文典型错误描述。"""
    kps = (q.get("qwen_analysis") or {}).get("knowledge_points") or []
    kp_str = "、".join(kps[:2]) if kps else "相关知识点"
    qtype = q.get("question_type")
    if qtype == "single_choice":
        return f"混淆{kp_str}的概念辨析；选项干扰项设计迷惑"
    if qtype == "blank":
        return f"未掌握{kp_str}的应用条件；书写不规范导致扣分"
    # solution
    return f"解题过程缺少对{kp_str}的关键步骤；分类讨论不完整"


def _stats(scores: list[int], full_score: int) -> dict:
    n = len(scores)
    avg = sum(scores) / n
    var = sum((s - avg) ** 2 for s in scores) / n
    std = var ** 0.5
    zeros = sum(1 for s in scores if s == 0) / n
    fulls = sum(1 for s in scores if s == full_score) / n
    err_rate = sum(1 for s in scores if s < full_score) / n
    return {
        "avg": round(avg, 2),
        "max": max(scores),
        "min": min(scores),
        "std": round(std, 2),
        "err_rate": round(err_rate, 4),
        "zero_rate": round(zeros, 4),
        "full_rate": round(fulls, 4),
    }


def build_workbook(paper_dir: Path, out_path: Path) -> dict:
    paper = json.loads((paper_dir / "paper.json").read_text(encoding="utf-8"))
    questions = paper["questions"]
    total = len(questions)

    wb = Workbook()
    ws = wb.active
    ws.title = "班级阅卷统计"

    # ---- 标题 + 元信息
    ws["A1"] = f"{paper['source']['name']} · 50 人班级阅卷统计"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADER))

    ws["A2"] = f"参考人数：{N_STUDENTS}　·　总分：{sum(q['full_score'] for q in questions)} 分"
    ws["A2"].font = Font(size=11, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADER))

    # ---- 表头
    for col, name in enumerate(HEADER, 1):
        c = ws.cell(row=4, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2C5FC7")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ---- 每题聚合
    for i, q in enumerate(questions):
        row = 5 + i
        qtype = q["question_type"]
        full = q["full_score"]
        # question_no 可能是 "1" 或 "二、(1)" 等，提取首个数字
        import re
        m = re.search(r"\d+", str(q["question_no"]))
        qno_int = int(m.group()) if m else (i + 1)
        diff = _classify_difficulty(qno_int, qtype, total)
        err_target = _target_error_rate(qtype, diff)

        scores = _gen_student_scores(N_STUDENTS, full, err_target, diff)
        s = _stats(scores, full)
        typical = _typical_error(q)

        ws.cell(row=row, column=1,  value=str(q["question_no"]).strip())
        ws.cell(row=row, column=2,  value={"single_choice": "单选", "multiple_choice": "多选", "blank": "填空", "solution": "解答"}.get(qtype, qtype))
        ws.cell(row=row, column=3,  value=full)
        ws.cell(row=row, column=4,  value=s["avg"])
        ws.cell(row=row, column=5,  value=s["max"])
        ws.cell(row=row, column=6,  value=s["min"])
        ws.cell(row=row, column=7,  value=s["std"])
        ws.cell(row=row, column=8,  value=s["err_rate"])
        ws.cell(row=row, column=9,  value=s["zero_rate"])
        ws.cell(row=row, column=10, value=s["full_rate"])
        ws.cell(row=row, column=11, value=typical)
        ws.cell(row=row, column=12, value=diff)

        # 错题率列百分比样式 + 条件底色
        for col, val in [(4, s["avg"]), (7, s["std"])]:
            ws.cell(row=row, column=col).number_format = "0.00"
        for col in (5, 6):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        for col in (8, 9, 10):
            cell = ws.cell(row=row, column=col)
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="center")
            # 错题率 > 50% 红，> 30% 黄，否则白
            if val >= 0.5:
                cell.fill = PatternFill("solid", fgColor="F8D7DA")
            elif val >= 0.3:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")

    # ---- 列宽
    widths = [10, 8, 8, 12, 10, 10, 10, 12, 12, 12, 50, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- 班级汇总
    sum_row = 5 + total + 1
    ws.cell(row=sum_row, column=1, value="班级汇总").font = Font(bold=True)
    sum_avg = 0.0
    for qi, q in enumerate(questions):
        qtype = q["question_type"]
        import re as _re
        m = _re.search(r"\d+", str(q["question_no"]))
        qno_int = int(m.group()) if m else (qi + 1)
        diff = _classify_difficulty(qno_int, qtype, total)
        err_target = _target_error_rate(qtype, diff)
        scores = _gen_student_scores(N_STUDENTS, q["full_score"], err_target, diff)
        sum_avg += _stats(scores, q["full_score"])["avg"]
    avg_total = round(sum_avg, 2)
    ws.cell(row=sum_row, column=4, value=avg_total)
    ws.cell(row=sum_row, column=4).font = Font(bold=True)

    # 及格率/优秀率行（按总分推算，但标注"仅供参考"）
    # 这里需要逐人总分，先聚合
    # 简化：用每题均分求和得到班级均分；分布按正态近似估算及格/优秀人数
    import statistics
    # 用每题学生得分来估算全班总分
    pass_row = sum_row + 1
    exc_row = sum_row + 2

    # 计算每位学生的总分
    student_totals = [0.0] * N_STUDENTS
    for qi, q in enumerate(questions):
        qtype = q["question_type"]
        m = re.search(r"\d+", str(q["question_no"]))
        qno_int = int(m.group()) if m else (qi + 1)
        diff = _classify_difficulty(qno_int, qtype, total)
        err_target = _target_error_rate(qtype, diff)
        scores = _gen_student_scores(N_STUDENTS, q["full_score"], err_target, diff)
        for si, s in enumerate(scores):
            student_totals[si] += s
    total_full = sum(q["full_score"] for q in questions)
    n_pass = sum(1 for s in student_totals if s >= total_full * 0.6)
    n_excellent = sum(1 for s in student_totals if s >= total_full * 0.85)
    avg_student_total = round(sum(student_totals) / N_STUDENTS, 2)
    std_student_total = round(statistics.pstdev(student_totals), 2)

    ws.cell(row=pass_row, column=1, value="班级均分（参考）").font = Font(italic=True, color="666666")
    ws.cell(row=pass_row, column=4, value=avg_student_total)
    ws.cell(row=pass_row, column=7, value=std_student_total)
    ws.cell(row=pass_row, column=8, value=f"{n_pass}/{N_STUDENTS} ({(n_pass/N_STUDENTS)*100:.1f}%) 及格")
    ws.cell(row=pass_row, column=10, value=f"{n_excellent}/{N_STUDENTS} ({(n_excellent/N_STUDENTS)*100:.1f}%) 优秀")

    # ---- 隐藏 sheet：班级汇总数据（schema 读取用）
    sum_ws = wb.create_sheet("班级汇总数据")
    sum_ws["A1"] = "key"; sum_ws["B1"] = "value"
    rows_meta = [
        ("n_students", N_STUDENTS),
        ("total_score", total_full),
        ("avg_total_score", avg_student_total),
        ("std_total_score", std_student_total),
        ("n_pass", n_pass),
        ("n_excellent", n_excellent),
        ("pass_rate", round(n_pass / N_STUDENTS, 4)),
        ("excellent_rate", round(n_excellent / N_STUDENTS, 4)),
    ]
    for i, (k, v) in enumerate(rows_meta, start=2):
        sum_ws.cell(row=i, column=1, value=k)
        sum_ws.cell(row=i, column=2, value=v)
    sum_ws.sheet_state = "hidden"

    wb.save(out_path)

    return {
        "paper": paper["source"]["name"],
        "out": str(out_path),
        "questions": total,
        "avg_total": avg_student_total,
        "n_pass": n_pass,
        "n_excellent": n_excellent,
    }


def main() -> None:
    out_dir = PAPERS_DIR / "_阅卷统计"
    out_dir.mkdir(exist_ok=True)
    summaries = []
    for d in sorted(PAPERS_DIR.iterdir()):
        if not d.is_dir() or d.name.startswith("_"):
            continue
        if not (d / "paper.json").exists():
            continue
        out = out_dir / f"{d.name}_班级阅卷统计.xlsx"
        info = build_workbook(d, out)
        info["out"] = str(out.relative_to(PAPERS_DIR.parent))
        summaries.append(info)
        print(json.dumps(info, ensure_ascii=False, indent=2))

    print("\n--- 汇总 ---")
    for s in summaries:
        print(f"  {s['paper']}: {s['questions']} 题, 班级总分均值 {s['avg_total']} -> {s['out']}")


if __name__ == "__main__":
    main()
