"""教学反馈模块：解析教师上传的 Excel -> 每题统计 + 班级汇总。"""
from __future__ import annotations

from dataclasses import asdict, dataclass, field
from typing import Any

from openpyxl import load_workbook


# 列名归一化：表头可能有"班级均分 / 平均分 / 均分"等同义表达
COLUMN_ALIASES: dict[str, list[str]] = {
    "no":           ["题号", "题序", "no", "q_no"],
    "type":         ["题型", "类型", "qtype"],
    "full_score":   ["分值", "满分", "full_score"],
    "avg_score":    ["班级均分", "平均分", "均分", "avg"],
    "max_score":    ["最高分", "max"],
    "min_score":    ["最低分", "min"],
    "std":          ["标准差", "方差", "std"],
    "error_rate":   ["错题率", "错误率", "err_rate"],
    "zero_rate":    ["零分率", "zero_rate"],
    "full_rate":    ["满分率", "full_rate"],
    "typical_error": ["典型错误", "高频错误", "典型错因"],
    "difficulty":   ["难度等级", "难度", "difficulty"],
}


@dataclass
class QuestionStat:
    no: str                       # "1" / "二、(1)" 等
    qtype: str = ""               # 中文：单选 / 填空 / 解答
    full_score: float = 0.0
    avg_score: float = 0.0
    max_score: float = 0.0
    min_score: float = 0.0
    std: float = 0.0
    error_rate: float = 0.0
    zero_rate: float = 0.0
    full_rate: float = 0.0
    typical_error: str = ""
    difficulty: str = ""          # 易 / 中 / 难

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class ClassReport:
    paper_id: str
    paper_name: str
    n_students: int = 0
    total_score: float = 0.0
    avg_total_score: float = 0.0
    pass_rate: float = 0.0          # >= 60%
    excellent_rate: float = 0.0     # >= 85%
    questions: list[QuestionStat] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        d["questions"] = [q.to_dict() for q in self.questions]
        return d


def _norm_header(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "")


def _map_headers(header_row: list[str]) -> dict[str, int]:
    """返回 canonical key -> 列索引（0-based）"""
    norm = [_norm_header(h) for h in header_row]
    mapping: dict[str, int] = {}
    for key, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            aliased = _norm_header(alias)
            if aliased in norm:
                mapping[key] = norm.index(aliased)
                break
    return mapping


def _to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().rstrip("%")
    try:
        return float(s)
    except ValueError:
        return 0.0


import re as _re
_NO_PREFIX_RE = _re.compile(r"^第\s*([\dA-Za-z（）()一二三四五六七八九十]+)\s*题\s*$")


def _normalize_no(s: str) -> str:
    """归一化题号：剥掉"第 X 题"前缀，只留内部 token。

    Excel 里教师可能写 "1" / "第 1 题" / "第1题" 都统一为 "1"；
    模板里再统一加 "第 X 题" 渲染。
    """
    m = _NO_PREFIX_RE.match(s)
    if m:
        return m.group(1)
    return s


def _norm_pct(v: float, col: str) -> float:
    """错题率/零分率/满分率：如果是 0~1 当小数；如果是 1~100 当百分比。"""
    if col in ("error_rate", "zero_rate", "full_rate") and v > 1.0:
        return v / 100.0
    return v


def parse_excel(path: str) -> tuple[list[QuestionStat], dict[str, float]]:
    """解析教师上传的 Excel（每题一行）。

    约定：第 1 行标题、第 2 行元信息、第 4 行表头、第 5 行起数据。
    但允许表头在任意一行——自动识别包含『题号』的最近一行。

    返回 (questions, class_meta)，class_meta 包含可选字段:
    n_students / n_pass / n_excellent / total_score / avg_total_score / pass_rate / excellent_rate
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # 找表头行
    header_row_idx = None
    mapping: dict[str, int] = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        m = _map_headers(list(row))
        if "no" in m:
            header_row_idx = row_idx
            mapping = m
            break
    if header_row_idx is None or "no" not in mapping:
        raise ValueError(f"Excel 未识别到'题号'列，请检查表头。当前表头：{[c.value for c in ws[1]]}")

    # 数据行
    stats: list[QuestionStat] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        no_val = row[mapping["no"]] if mapping["no"] < len(row) else None
        if no_val is None or str(no_val).strip() == "" or "汇总" in str(no_val):
            continue
        no = _normalize_no(str(no_val).strip())
        qtype = str(row[mapping.get("type", 0)] if "type" in mapping and mapping["type"] < len(row) else "").strip()
        full = _to_float(row[mapping["full_score"]] if mapping["full_score"] < len(row) else 0)
        avg = _to_float(row[mapping.get("avg_score", 0)] if "avg_score" in mapping and mapping["avg_score"] < len(row) else 0)
        mx = _to_float(row[mapping.get("max_score", 0)] if "max_score" in mapping and mapping["max_score"] < len(row) else 0)
        mn = _to_float(row[mapping.get("min_score", 0)] if "min_score" in mapping and mapping["min_score"] < len(row) else 0)
        std = _to_float(row[mapping.get("std", 0)] if "std" in mapping and mapping["std"] < len(row) else 0)
        er = _norm_pct(_to_float(row[mapping["error_rate"]] if mapping["error_rate"] < len(row) else 0), "error_rate")
        zr = _norm_pct(_to_float(row[mapping.get("zero_rate", 0)] if "zero_rate" in mapping and mapping["zero_rate"] < len(row) else 0), "zero_rate")
        fr = _norm_pct(_to_float(row[mapping.get("full_rate", 0)] if "full_rate" in mapping and mapping["full_rate"] < len(row) else 0), "full_rate")
        typerr = str(row[mapping.get("typical_error", 0)] if "typical_error" in mapping and mapping["typical_error"] < len(row) else "").strip()
        diff = str(row[mapping.get("difficulty", 0)] if "difficulty" in mapping and mapping["difficulty"] < len(row) else "").strip()

        stats.append(QuestionStat(
            no=no, qtype=qtype, full_score=full,
            avg_score=avg, max_score=mx, min_score=mn, std=std,
            error_rate=er, zero_rate=zr, full_rate=fr,
            typical_error=typerr, difficulty=diff,
        ))

    # 隐藏的"班级汇总数据" sheet：key/value
    meta: dict[str, float] = {}
    if "班级汇总数据" in wb.sheetnames:
        sum_ws = wb["班级汇总数据"]
        for r in sum_ws.iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                meta[str(r[0])] = _to_float(r[1]) if len(r) > 1 else 0.0
    return stats, meta


def build_class_report(
    paper_id: str,
    paper_name: str,
    n_students: int,
    questions: list[QuestionStat],
    n_pass: int | None = None,
    n_excellent: int | None = None,
    total_score: float | None = None,
    avg_total_score: float | None = None,
    pass_rate: float | None = None,
    excellent_rate: float | None = None,
) -> ClassReport:
    """构造班级层面的报告对象。"""
    total = total_score if total_score is not None else sum(q.full_score for q in questions)
    avg_total = round(avg_total_score if avg_total_score is not None else sum(q.avg_score for q in questions), 2)
    if pass_rate is not None:
        pr = pass_rate
    elif n_pass is not None and n_students:
        pr = n_pass / n_students
    else:
        pr = 0.0
    if excellent_rate is not None:
        er = excellent_rate
    elif n_excellent is not None and n_students:
        er = n_excellent / n_students
    else:
        er = 0.0
    return ClassReport(
        paper_id=paper_id,
        paper_name=paper_name,
        n_students=n_students,
        total_score=total,
        avg_total_score=avg_total,
        pass_rate=round(pr, 4),
        excellent_rate=round(er, 4),
        questions=questions,
    )


def weak_questions(report: ClassReport, top_k: int = 5) -> list[QuestionStat]:
    """按错题率排序，取 top_k 薄弱题。"""
    return sorted(report.questions, key=lambda q: q.error_rate, reverse=True)[:top_k]


def qtype_aggregate(report: ClassReport) -> list[dict[str, Any]]:
    """按题型聚合错题率。"""
    buckets: dict[str, list[QuestionStat]] = {}
    for q in report.questions:
        buckets.setdefault(q.qtype or "其他", []).append(q)
    rows = []
    for qt, items in buckets.items():
        if not items:
            continue
        avg_er = sum(q.error_rate for q in items) / len(items)
        avg_score = sum(q.avg_score for q in items) / len(items)
        rows.append({
            "qtype": qt,
            "count": len(items),
            "avg_error_rate": round(avg_er, 4),
            "avg_score": round(avg_score, 2),
        })
    rows.sort(key=lambda r: r["avg_error_rate"], reverse=True)
    return rows
